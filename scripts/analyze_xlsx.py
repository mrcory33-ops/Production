import openpyxl
import os
import json
from collections import Counter

filepath = r"C:\Users\CoryD\OneDrive - Emjac Industries, Inc\Desktop\#9's.xlsx"
print(f"File exists: {os.path.exists(filepath)}")
print(f"File size: {os.path.getsize(filepath)} bytes")

wb = openpyxl.load_workbook(filepath, data_only=True)
print(f"\nSheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Find the header row (first non-empty row)
    header_row = None
    for row_idx in range(1, min(10, ws.max_row + 1)):
        row_vals = [cell.value for cell in ws[row_idx]]
        non_empty = [v for v in row_vals if v is not None]
        if len(non_empty) > 3:
            header_row = row_idx
            break
    
    if header_row is None:
        print("  Could not find header row")
        continue
    
    print(f"\nHeader Row: {header_row}")
    headers = []
    for i, cell in enumerate(ws[header_row], 1):
        if cell.value is not None:
            headers.append((i, str(cell.value)))
            print(f"  Col {i}: {cell.value}")
    
    # Show sample data rows
    print(f"\n--- Sample Data (first 10 data rows) ---")
    data_start = header_row + 1
    row_count = 0
    for row_idx in range(data_start, min(data_start + 10, ws.max_row + 1)):
        row_vals = [cell.value for cell in ws[row_idx]]
        if any(v is not None for v in row_vals):
            row_count += 1
            print(f"  Row {row_idx}: {row_vals}")
    
    # Analyze data types per column
    print(f"\n--- Column Data Type Analysis ---")
    for col_idx, col_name in headers:
        types = Counter()
        sample_vals = []
        for row_idx in range(data_start, min(data_start + 100, ws.max_row + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                types[type(val).__name__] += 1
                if len(sample_vals) < 5:
                    sample_vals.append(str(val)[:50])
        print(f"  {col_name}: types={dict(types)}, samples={sample_vals}")
    
    # Count actual data rows (non-empty)
    actual_rows = 0
    for row_idx in range(data_start, ws.max_row + 1):
        row_vals = [cell.value for cell in ws[row_idx]]
        if any(v is not None for v in row_vals):
            actual_rows += 1
    print(f"\nTotal data rows (non-empty): {actual_rows}")

print("\n\nDONE")
