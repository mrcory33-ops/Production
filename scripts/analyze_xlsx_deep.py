import openpyxl
from collections import Counter, defaultdict

filepath = r"C:\Users\CoryD\OneDrive - Emjac Industries, Inc\Desktop\#9's.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb['JCS']

# Understand the hierarchical structure
# Let's categorize each row type
print("=== ROW TYPE ANALYSIS ===")
customer_rows = 0
job_rows = 0
component_rows = 0
other_rows = 0

jobs = set()
customers = set()
sales_orders = set()
components = set()
vendors = set()
po_numbers = set()
code_sorts = Counter()

# Track jobs and their components
job_components = defaultdict(list)
current_customer = None
current_job = None
current_mark = None

for row_idx in range(2, ws.max_row + 1):
    row = {
        'Text48': ws.cell(row=row_idx, column=1).value,
        'Auto_Date': ws.cell(row=row_idx, column=2).value,
        'SALES_ORDER': ws.cell(row=row_idx, column=3).value,
        'MARK_INFO': ws.cell(row=row_idx, column=4).value,
        'SALESPERSON': ws.cell(row=row_idx, column=5).value,
        'CUSTOMER': ws.cell(row=row_idx, column=6).value,
        'CODE_SORT': ws.cell(row=row_idx, column=7).value,
        'JOB': ws.cell(row=row_idx, column=8).value,
        'PART_CUSTOMER': ws.cell(row=row_idx, column=9).value,
        'Text72': ws.cell(row=row_idx, column=10).value,
        'COMPONENT': ws.cell(row=row_idx, column=11).value,
        'DESCRIPTION': ws.cell(row=row_idx, column=12).value,
        'UM': ws.cell(row=row_idx, column=13).value,
        'QTY_COMMITTED': ws.cell(row=row_idx, column=14).value,
        'QTY_ISSUED': ws.cell(row=row_idx, column=15).value,
        'QTY_ONHAND': ws.cell(row=row_idx, column=16).value,
        'PURCHASE_ORDER': ws.cell(row=row_idx, column=17).value,
        'VENDOR': ws.cell(row=row_idx, column=18).value,
        'QTY_ORDER': ws.cell(row=row_idx, column=19).value,
        'QTY_RECEIVED': ws.cell(row=row_idx, column=20).value,
        'DATE_DUE_LINE': ws.cell(row=row_idx, column=21).value,
        'DATE_LAST_RECEIVED': ws.cell(row=row_idx, column=22).value,
    }
    
    if row['MARK_INFO'] and row['CUSTOMER']:
        # This is a customer/project header row
        customer_rows += 1
        current_customer = row['CUSTOMER']
        current_mark = row['MARK_INFO']
        customers.add(current_customer)
        if row['SALES_ORDER']:
            sales_orders.add(row['SALES_ORDER'])
        if row['CODE_SORT']:
            code_sorts[row['CODE_SORT']] += 1
    elif row['JOB']:
        # This is a job row
        job_rows += 1
        current_job = row['JOB']
        jobs.add(current_job)
    elif row['COMPONENT']:
        # This is a component detail row
        component_rows += 1
        components.add(row['COMPONENT'])
        if row['VENDOR']:
            vendors.add(row['VENDOR'])
        if row['PURCHASE_ORDER']:
            po_numbers.add(row['PURCHASE_ORDER'])
        
        if current_job:
            has_po = row['PURCHASE_ORDER'] is not None
            qty_committed = row['QTY_COMMITTED'] or 0
            qty_onhand = row['QTY_ONHAND'] or 0
            qty_issued = row['QTY_ISSUED'] or 0
            shortage = qty_committed - qty_onhand if qty_committed > qty_onhand else 0
            
            job_components[current_job].append({
                'component': row['COMPONENT'],
                'description': row['DESCRIPTION'],
                'qty_committed': qty_committed,
                'qty_issued': qty_issued,
                'qty_onhand': qty_onhand,
                'has_po': has_po,
                'po': row['PURCHASE_ORDER'],
                'vendor': row['VENDOR'],
                'qty_ordered': row['QTY_ORDER'],
                'qty_received': row['QTY_RECEIVED'],
                'date_due': str(row['DATE_DUE_LINE']) if row['DATE_DUE_LINE'] else None,
                'shortage': shortage,
            })
    else:
        other_rows += 1

print(f"Customer/Project header rows: {customer_rows}")
print(f"Job rows: {job_rows}")
print(f"Component detail rows: {component_rows}")
print(f"Other/empty rows: {other_rows}")

print(f"\n=== UNIQUE COUNTS ===")
print(f"Unique customers: {len(customers)}")
print(f"Unique jobs: {len(jobs)}")
print(f"Unique sales orders: {len(sales_orders)}")
print(f"Unique components: {len(components)}")
print(f"Unique vendors: {len(vendors)}")
print(f"Unique PO numbers: {len(po_numbers)}")

print(f"\n=== CODE_SORT VALUES ===")
for code, count in code_sorts.most_common():
    print(f"  {code}: {count}")

print(f"\n=== CUSTOMER LIST ===")
for c in sorted(customers):
    print(f"  {c}")

print(f"\n=== SALES ORDER LIST ===")
for so in sorted(sales_orders):
    print(f"  {so}")

print(f"\n=== JOB LIST ===")
for j in sorted(jobs):
    print(f"  {j}")

# Analyze component shortages and PO status
print(f"\n=== PURCHASE ORDER ANALYSIS ===")
total_components_with_po = 0
total_components_stock = 0
po_received_complete = 0
po_partially_received = 0
po_not_received = 0

for job, comps in job_components.items():
    for comp in comps:
        if comp['has_po']:
            total_components_with_po += 1
            qty_ord = comp['qty_ordered'] or 0
            qty_rec = comp['qty_received'] or 0
            if qty_rec >= qty_ord and qty_ord > 0:
                po_received_complete += 1
            elif qty_rec > 0:
                po_partially_received += 1
            else:
                po_not_received += 1
        else:
            total_components_stock += 0

print(f"Components with PO (#9 purchases): {total_components_with_po}")
print(f"PO fully received: {po_received_complete}")
print(f"PO partially received: {po_partially_received}")
print(f"PO not yet received: {po_not_received}")

# Analyze stock shortages
print(f"\n=== STOCK AVAILABILITY ANALYSIS ===")
stock_sufficient = 0
stock_shortage = 0
for job, comps in job_components.items():
    for comp in comps:
        if not comp['has_po']:
            if comp['qty_onhand'] >= comp['qty_committed']:
                stock_sufficient += 1
            else:
                stock_shortage += 1

print(f"Stock items with sufficient qty: {stock_sufficient}")
print(f"Stock items with shortage: {stock_shortage}")

# Show a few example jobs with their full component breakdown
print(f"\n=== EXAMPLE JOB BREAKDOWNS (first 3 jobs with POs) ===")
shown = 0
for job in sorted(job_components.keys()):
    comps = job_components[job]
    has_any_po = any(c['has_po'] for c in comps)
    if has_any_po and shown < 3:
        shown += 1
        print(f"\n  JOB: {job}")
        for c in comps:
            po_info = f"PO#{c['po']} vendor:{c['vendor']} ordered:{c['qty_ordered']} rcvd:{c['qty_received']} due:{c['date_due']}" if c['has_po'] else "STOCK"
            print(f"    {c['component']} - {c['description']}: committed={c['qty_committed']} onhand={c['qty_onhand']} issued={c['qty_issued']} | {po_info}")
